import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

def create_excel_template():
    """Excel template'i oluştur - 80 paket ile"""
    
    # Yeni workbook oluştur
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Varsayılan sheet'i sil
    
    # Sayfaları oluştur
    ws_araclar = wb.create_sheet("Araçlar")
    ws_paketler = wb.create_sheet("Paketler")
    ws_sonuclar = wb.create_sheet("Sonuçlar")
    
    # Stil tanımlamaları
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # ==================== ARAÇLAR SAYFASI ====================
    print("Araçlar sayfası oluşturuluyor...")
    
    headers_araclar = ["ID", "Uzunluk (cm)", "Genişlik (cm)", "Yükseklik (cm)", "Max Ağırlık (kg)"]
    
    for col, header in enumerate(headers_araclar, 1):
        cell = ws_araclar.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Araç verisi
    araclar = [
        [1, 600, 250, 250, 25000],
        [2, 800, 250, 280, 30000],
        [3, 1000, 300, 300, 40000]
    ]
    
    for row_idx, arac in enumerate(araclar, 2):
        for col_idx, value in enumerate(arac, 1):
            cell = ws_araclar.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = center_alignment
    
    ws_araclar.column_dimensions['A'].width = 12
    ws_araclar.column_dimensions['B'].width = 18
    ws_araclar.column_dimensions['C'].width = 18
    ws_araclar.column_dimensions['D'].width = 18
    ws_araclar.column_dimensions['E'].width = 20
    
    # ==================== PAKETLER SAYFASI ====================
    print("Paketler sayfası oluşturuluyor...")
    
    headers_paketler = ["ID", "Uzunluk (cm)", "Genişlik (cm)", "Yükseklik (cm)", "Ağırlık (kg)"]
    
    for col, header in enumerate(headers_paketler, 1):
        cell = ws_paketler.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # 80 paket verisi oluştur
    # Ağır paketler (Paket 1-20): 15-25 kg
    # Orta paketler (Paket 21-50): 8-14 kg
    # Hafif paketler (Paket 51-80): 2-7 kg
    
    paketler = []
    
    # Ağır paketler
    for i in range(1, 21):
        uzunluk = random.randint(100, 150)
        genislik = random.randint(60, 100)
        yukseklik = random.randint(50, 90)
        agirlik = random.randint(15, 25)
        paketler.append([i, uzunluk, genislik, yukseklik, agirlik])
    
    # Orta paketler
    for i in range(21, 51):
        uzunluk = random.randint(80, 130)
        genislik = random.randint(50, 90)
        yukseklik = random.randint(40, 80)
        agirlik = random.randint(8, 14)
        paketler.append([i, uzunluk, genislik, yukseklik, agirlik])
    
    # Hafif paketler
    for i in range(51, 81):
        uzunluk = random.randint(60, 110)
        genislik = random.randint(40, 70)
        yukseklik = random.randint(30, 60)
        agirlik = random.randint(2, 7)
        paketler.append([i, uzunluk, genislik, yukseklik, agirlik])
    
    for row_idx, paket in enumerate(paketler, 2):
        for col_idx, value in enumerate(paket, 1):
            cell = ws_paketler.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = center_alignment
    
    ws_paketler.column_dimensions['A'].width = 12
    ws_paketler.column_dimensions['B'].width = 18
    ws_paketler.column_dimensions['C'].width = 18
    ws_paketler.column_dimensions['D'].width = 18
    ws_paketler.column_dimensions['E'].width = 16
    
    # ==================== SONUÇLAR SAYFASI ====================
    print("Sonuçlar sayfası oluşturuluyor...")
    
    ws_sonuclar.cell(row=1, column=1).value = "YÜKLEME OPTİMİZASYONU SONUÇLARI"
    ws_sonuclar.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    # İstatistikler alanı
    ws_sonuclar.cell(row=3, column=1).value = "Seçili Araç ID:"
    ws_sonuclar.cell(row=4, column=1).value = "Toplam Ağırlık (kg):"
    ws_sonuclar.cell(row=5, column=1).value = "Kapasite (kg):"
    ws_sonuclar.cell(row=6, column=1).value = "Kullanılan Hacim (%):"
    
    # Başlıklar
    headers_sonuclar = ["Paket ID", "X (cm)", "Y (cm)", "Z (cm)", "Uzunluk", "Genişlik", "Yükseklik", "Ağırlık", "Durum"]
    
    for col, header in enumerate(headers_sonuclar, 1):
        cell = ws_sonuclar.cell(row=8, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    ws_sonuclar.column_dimensions['A'].width = 12
    for col in range(2, 10):
        ws_sonuclar.column_dimensions[get_column_letter(col)].width = 14
    
    # Dosyayı kaydet
    file_path = "Yükleme_Optimizasyonu.xlsm"
    wb.save(file_path)
    print(f"✅ Excel dosyası başarıyla oluşturuldu: {file_path}")

if __name__ == "__main__":
    create_excel_template()
